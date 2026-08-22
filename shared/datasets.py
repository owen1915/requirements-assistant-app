"""Load PM and Bookshelf requirements + ground truth from the survey workbooks.

Everything is derived from the two "Ground Truth - *.xlsx" files so the testbed
is self-contained and reproducible. Each workbook has:

  * per-category sheets (Functional/Performance/Environment/Resource
    Requirements) — each requirement is a header row "<CAT> <n>: <text>",
    followed by ~11 issue rows carrying the expert's per-issue Y/N + Comment.
  * an "Aggregate" sheet — a 23 issue-type x requirement matrix, cell "x" where
    that issue applies. We collapse the issue rows to the 7 INCOSE criteria the
    prototype evaluates, via ISSUE_TO_CRITERION below.

Coverage note: the survey's issue taxonomy cleanly covers A2/A4/A5/A6/A9/A10.
It has no clean row for A3 (Appropriate) — so A3 has no GT signal and is
reported separately. Its "Verifiable"/"Feasible" rows map to A8/A7, which the
prototype does not evaluate, so they are excluded from the KPI comparison.
"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List

import openpyxl

# The two workbooks live one directory above the repo root (where they were
# downloaded). Resolve relative to this file so the path is stable.
from shared.paths import GT_DIR

WORKBOOKS = {
    "PM": GT_DIR / "Ground Truth - Good Requiremnts.xlsx",
    "Bookshelf": GT_DIR / "Ground Truth - Bad Requiremnts.xlsx",
}

# Preferred ground truth: SMEs annotated the prototype's criteria DIRECTLY
# (A2-A10), so no taxonomy mapping is needed and scoring aligns with what the
# evaluator reports. Stored in R-series; we invert the testbed's known A->R map.
# Only PM has this today; Bookshelf falls back to the survey aggregate.
ASERIES_GT_FILES = {"PM": GT_DIR / "R1_R14_GroundTruth.xlsx"}

# The v5 ground truth is annotated directly in A-criteria form (A1-A11 rows plus
# a B block and FR/PR/ER/RR/ALL set columns) - the same matrix shape the n8n
# workflow emits. Unlike the R-series and survey ground truths, it carries A2 and
# A3 positives, so those criteria are actually scorable against it.
V5_GT_FILES = {
    "PM": GT_DIR / "Point_Mass_v5_GroundTruth.xlsx",
    "Bookshelf": GT_DIR / "BookShelf_v5_GroundTruth.xlsx",
}
R_TO_A = {"R1": "A2", "R2": "A6", "R3": "A4", "R4": "A5",
          "R5": "A9", "R11": "A10", "R12": "A3"}

CATEGORY_SHEETS = {
    "Functional Requirements": "FR",
    "Performance Requirements": "PR",
    "Environment Requirements": "ER",
    "Resource Requirements": "RR",
}

# The 7 criteria the prototype evaluates (A-series). A3 kept for completeness.
CRITERIA = ["A2", "A3", "A4", "A5", "A6", "A9", "A10"]
CRITERIA_NAMES = {
    "A2": "Necessary", "A3": "Appropriate", "A4": "Unambiguous", "A5": "Complete",
    "A6": "Singular", "A9": "Correct", "A10": "Conforming",
}

# Aggregate issue-row (0-indexed within the Issue column) -> prototype criterion.
# Rows 8 (Verifiable=A8) and 9 (Feasible=A7) are intentionally unmapped: the
# prototype does not evaluate them, so they cannot enter a fair KPI comparison.
ISSUE_TO_CRITERION = {
    1: "A10",   # Improper Format / Missing "Shall"
    2: "A4",    # Lack of Ambiguity
    3: "A4",    # Clarity
    4: "A9",    # Correctness
    5: "A6",    # Conciseness (define only one thing)
    6: "A2",    # Necessity
    7: "A5",    # Properly Bounded
}

_HEADER_RE = re.compile(r"^(FR|PR|ER|RR)\s*0*(\d+)\s*:\s*(.*)", re.IGNORECASE | re.DOTALL)

_MOJIBAKE = {"â€\x9d": '"', "â€œ": '"', "â€™": "'", "â€“": "-", "�": '"', "â": '"'}


def _clean(text: str) -> str:
    text = text or ""
    for bad, good in _MOJIBAKE.items():
        text = text.replace(bad, good)
    return " ".join(text.split()).strip()


@dataclass
class Requirement:
    req_id: str          # "FR.1", "PR.3", ...  (matches GT/aggregate columns)
    category: str        # FR | PR | ER | RR
    text: str
    # expert per-issue comments for this requirement (issue -> comment), for NLP
    comments: Dict[str, str] = field(default_factory=dict)


@dataclass
class Dataset:
    name: str
    requirements: List[Requirement]
    gt: Dict[str, Dict[str, bool]]   # criterion -> {req_id -> violated?}

    @property
    def req_ids(self) -> List[str]:
        return [r.req_id for r in self.requirements]


def _load_requirements(wb) -> List[Requirement]:
    reqs: List[Requirement] = []
    for sheet_name, prefix in CATEGORY_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        current: Requirement | None = None
        for r in range(1, ws.max_row + 1):
            a = ws.cell(r, 1).value
            if a is None:
                continue
            m = _HEADER_RE.match(str(a).strip())
            if m and m.group(1).upper() == prefix:
                req_id = f"{prefix}.{int(m.group(2))}"
                current = Requirement(req_id=req_id, category=prefix,
                                      text=_clean(m.group(3)))
                reqs.append(current)
            elif current is not None:
                # issue row under the current requirement: (issue, Y/N, comment)
                issue = _clean(str(a))
                comment = ws.cell(r, 3).value
                if comment and str(comment).strip():
                    current.comments[issue[:60]] = _clean(str(comment))
    return reqs


def _load_gt(wb, req_ids: List[str]) -> Dict[str, Dict[str, bool]]:
    ws = wb["Aggregate"]
    # Map aggregate header columns (row containing FR.1, PR.1, ...) to col index
    header_row = None
    for r in range(1, 4):
        vals = [str(ws.cell(r, c).value).strip() if ws.cell(r, c).value else ""
                for c in range(1, ws.max_column + 1)]
        if any(v in req_ids for v in vals):
            header_row = r
            break
    if header_row is None:
        raise ValueError("Could not locate the aggregate header row")

    col_of = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v and str(v).strip() in req_ids:
            col_of[str(v).strip()] = c

    gt = {crit: {rid: False for rid in req_ids} for crit in CRITERIA}
    # Issue rows begin right after the header row; index them from 0.
    issue_idx = 0
    for r in range(header_row + 1, ws.max_row + 1):
        crit = ISSUE_TO_CRITERION.get(issue_idx)
        if crit:
            for rid, c in col_of.items():
                cell = ws.cell(r, c).value
                if cell and str(cell).strip().lower() == "x":
                    gt[crit][rid] = True
        issue_idx += 1
    return gt


def _load_aseries_gt(path: Path, req_ids: List[str]) -> Dict[str, Dict[str, bool]]:
    """Load a directly-annotated A-criteria GT from an R-series workbook."""
    import pandas as pd
    df = pd.read_excel(path, sheet_name="Sheet1")
    gt = {crit: {rid: False for rid in req_ids} for crit in CRITERIA}
    for _, row in df.iterrows():
        crit = R_TO_A.get(str(row["Rule ID"]).strip())
        if not crit:
            continue
        for rid in req_ids:
            if rid in df.columns and str(row[rid]).strip().lower() == "x":
                gt[crit][rid] = True
    return gt


def _load_v5_gt(path: Path, req_ids: List[str]) -> Dict[str, Dict[str, bool]]:
    """Load a v5 ground truth: A-criteria rows keyed by 'Issue ID'.

    The sheet also carries a B block below a blank spacer row and FR/PR/ER/RR/ALL
    set columns to its right; both are ignored here, since this is the
    individual-criteria ground truth.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    col_of = {str(v).strip(): i for i, v in enumerate(header)
              if v and str(v).strip() in req_ids}
    missing = [r for r in req_ids if r not in col_of]
    if missing:
        raise ValueError(f"{path.name}: no column for {missing}")

    gt = {crit: {rid: False for rid in req_ids} for crit in CRITERIA}
    for row in rows[1:]:
        rid_cell = row[0]
        crit = str(rid_cell).strip().upper() if rid_cell else ""
        if crit not in CRITERIA:
            continue                        # A1/A11, the blank row, the B block
        for rid, ci in col_of.items():
            cell = row[ci] if ci < len(row) else None
            if cell and str(cell).strip().lower() == "x":
                gt[crit][rid] = True
    return gt


def gt_source() -> str:
    """Which ground truth to score against.

    "auto"    - v5 where available, else A-series (R1_R14), else survey.
    "v5"      - force the v5 A-criteria workbook (the current reference GT).
    "aseries" - force the R1_R14 file.
    "survey"  - force the survey-aggregate mapping in the workbook itself.

    These disagree materially. On PM the v5 GT has A2 and A3 positives while the
    other two have none, and A9 is 9 positives in v5 against 1 in the R-series -
    so which one is in play has to be an explicit, recorded choice rather than a
    silent fallback.
    """
    return os.getenv("AI4RE_GT_SOURCE", "auto").strip().lower()


def load_dataset(name: str, source: str | None = None) -> Dataset:
    path = WORKBOOKS[name]
    wb = openpyxl.load_workbook(path, data_only=True)
    reqs = _load_requirements(wb)
    req_ids = [r.req_id for r in reqs]

    src = (source or gt_source())
    has_v5 = name in V5_GT_FILES and V5_GT_FILES[name].exists()
    has_aseries = name in ASERIES_GT_FILES and ASERIES_GT_FILES[name].exists()

    if src == "v5" or (src == "auto" and has_v5):
        if not has_v5:
            raise ValueError(f"{name}: no v5 ground truth file available")
        gt = _load_v5_gt(V5_GT_FILES[name], req_ids)
    elif src == "aseries" or (src == "auto" and has_aseries):
        if not has_aseries:
            raise ValueError(f"{name}: no A-series ground truth file available")
        gt = _load_aseries_gt(ASERIES_GT_FILES[name], req_ids)
    else:
        gt = _load_gt(wb, req_ids)          # survey-aggregate mapping
    return Dataset(name=name, requirements=reqs, gt=gt)


def _main() -> None:
    for name in WORKBOOKS:
        ds = load_dataset(name)
        print(f"\n=== {name}: {len(ds.requirements)} requirements ===")
        for req in ds.requirements:
            viols = [c for c in CRITERIA if ds.gt[c][req.req_id]]
            print(f"  {req.req_id:6s} [{','.join(viols) or 'clean':<18}] "
                  f"{req.text[:60]}")
        print(f"  GT violations per criterion: "
              + ", ".join(f"{c}={sum(ds.gt[c].values())}" for c in CRITERIA))
        n_comments = sum(len(r.comments) for r in ds.requirements)
        print(f"  expert comments captured: {n_comments}")


if __name__ == "__main__":
    _main()
