"""Convert a real n8n workflow export into the testbed's matrix format.

The n8n workbook records the 11-row SURVEY issue taxonomy; the testbed records
A-criteria directly. Converting once, here, means every downstream tool —
metrics.py, boxplots, figures.py, the parity tests — reads the real n8n run with
no changes, and both arms are scored against the same ground truth by the same
code. That is the whole point: a second scoring path would be a second set of
assumptions to audit.

Two things the mapping has to get right:

  * A4 receives TWO n8n rows ("Lack of Ambiguity" and "Clarity"). They are OR'd
    — either one flags the criterion.
  * A3 has NO n8n row. It is left out of the output entirely rather than written
    blank, because a blank row scores as "never flagged" instead of "not
    evaluated" and would turn A3's ground-truth positives into false negatives.
    Score both arms with `criteria=COMPARABLE` so the pooling matches.

    python -m n8n_testbed.import_n8n "path/to/export.xlsx" --config n8nreal
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Dict, List

import openpyxl

from shared.datasets import CRITERIA_NAMES
from shared.paths import TESTBED_RUNS_DIR

# n8n survey issue id -> prototype criterion.
ISSUE_TO_CRITERION: Dict[int, str] = {
    2: "A10",   # Improper Requirement Format: Missing "Shall"
    3: "A4",    # Lack of Ambiguity
    4: "A4",    # Clarity            <- second row into the same criterion
    5: "A9",    # Correctness
    6: "A6",    # Conciseness
    7: "A2",    # Necessity
    8: "A5",    # Properly Bounded
}

# Deliberately unmapped: 1 = "no issues" (A1), 9 = Verifiable (A8),
# 10 = Feasible (A7), 11 = Unsure (A11). None is scored by the prototype.
UNMAPPED = {1: "A1 no-issues", 9: "A8 verifiable", 10: "A7 feasible", 11: "A11 unsure"}

# The criteria both sides actually evaluate. Pass to metrics.score(criteria=...).
COMPARABLE: List[str] = sorted(set(ISSUE_TO_CRITERION.values()))

_EXEC = re.compile(r"(\d+)")


def read_export(path: Path) -> Dict[str, Dict[str, Dict[str, bool]]]:
    """{sheet: {criterion: {req_id: flagged}}} for every populated sheet."""
    wb = openpyxl.load_workbook(path, data_only=True)
    out: Dict[str, Dict[str, Dict[str, bool]]] = {}
    skipped = []

    for name in wb.sheetnames:
        rows = [r for r in wb[name].iter_rows(values_only=True) if r and any(r)]
        if len(rows) < 3:
            skipped.append(name)          # an empty tab, not a run
            continue
        req_ids = [str(c).strip() for c in rows[0][2:] if c and str(c).strip()
                   and str(c).strip().lower() != "executionnumber"]

        flags = {c: {r: False for r in req_ids} for c in COMPARABLE}
        for row in rows[1:]:
            try:
                issue = int(row[0])
            except (TypeError, ValueError):
                continue                   # "Evaluation Result" spacer row
            crit = ISSUE_TO_CRITERION.get(issue)
            if not crit:
                continue
            for i, rid in enumerate(req_ids):
                cell = row[2 + i] if 2 + i < len(row) else None
                if cell and str(cell).strip().lower() == "x":
                    flags[crit][rid] = True      # OR: A4's two rows both land here
        out[name] = flags

    if skipped:
        print(f"  skipped {len(skipped)} empty sheet(s): "
              f"{skipped[0]} .. {skipped[-1]}")
    return out


def write_matrix(runs: Dict[str, Dict[str, Dict[str, bool]]], dest: Path) -> Path:
    """Write the testbed's Execution_NN matrix shape."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ordered = sorted(runs, key=lambda n: int(m.group(1))
                     if (m := _EXEC.search(n)) else 0)
    for i, name in enumerate(ordered, start=1):
        flags = runs[name]
        req_ids = list(next(iter(flags.values())))
        ws = wb.create_sheet(f"Execution_{i:02d}")
        ws.append(["Rule ID", "Rule Description"] + req_ids)
        for crit in COMPARABLE:
            ws.append([crit, CRITERIA_NAMES.get(crit, "")] +
                      ["x" if flags[crit][r] else None for r in req_ids])
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return dest


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("export", type=Path, help="the n8n .xlsx export")
    ap.add_argument("--dataset", default="PM")
    ap.add_argument("--config", default="n8nreal",
                    help="tag for the written matrix: PM_<config>_matrix.xlsx")
    a = ap.parse_args()

    runs = read_export(a.export)
    if not runs:
        raise SystemExit("no populated sheets in that workbook")
    dest = TESTBED_RUNS_DIR / f"{a.dataset}_{a.config}_matrix.xlsx"
    write_matrix(runs, dest)

    n_flags = sum(sum(v.values()) for r in runs.values() for v in r.values())
    print(f"  {len(runs)} runs -> {dest.name}")
    print(f"  criteria: {', '.join(COMPARABLE)}   (A3 omitted - no n8n row)")
    print(f"  {n_flags} flags total, {n_flags / len(runs):.1f} per run")
    print(f"\n  score with: metrics.score('{a.dataset}', '{a.config}', "
          f"criteria=import_n8n.COMPARABLE)")


if __name__ == "__main__":
    main()
