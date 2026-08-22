"""Run harness — produce prediction matrices for each dataset x config.

For every (dataset, config) it runs the evaluator N times over all requirements
and writes one workbook with an Execution_01..NN sheet per run, in the format the
KPI engine expects: rows = the 7 criteria, columns = requirement IDs, cell "x"
where the criterion was flagged as violated. Raw per-criterion evaluations are
also saved (JSON) for the NLP comment analysis.

    python -m n8n_testbed.run_matrix --runs 10 --datasets PM Bookshelf \
        --configs baseline M1 M2

Use --runs 1 --datasets PM --configs baseline for a quick smoke test.
"""

from __future__ import annotations

import argparse
import json
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Dict, List

import openpyxl

from shared import datasets, spec
from . import inject
from shared.datasets import CRITERIA, CRITERIA_NAMES

from shared.paths import TESTBED_RUNS_DIR as RUNS_DIR, output_dir

# Matrices, raw dumps and ground-truth workbooks are the RUN STORE: each one
# cost 100 whole-set LLM calls, so they are versioned rather than rebuilt.
# Plots and summary workbooks are regenerable and go to outputs/.
RUNS_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR = output_dir("n8n_testbed")

MAX_WORKERS = 6
MAX_RETRIES = 6

# Count of cells that exhausted retries and were recorded as no-flags, so a
# single flaky API response doesn't abort a multi-hour batch. Surfaced per run.
_DEGRADED = {"n": 0}


def _analyze_one(req_id, text, context, sme_rules, provider):
    """Analyze one requirement, retrying transient API/parse errors.

    If every retry fails (usually an intermittent empty response), record the
    requirement as no-flags rather than crashing the whole batch. These are
    counted and reported so the impact is visible, not hidden."""
    last = None
    for attempt in range(MAX_RETRIES):
        try:
            return inject.analyze_requirement(
                {"id": req_id, "text": text}, context, sme_rules, provider)
        except Exception as exc:  # noqa: BLE001
            last = exc
            time.sleep(2 * (attempt + 1))
    _DEGRADED["n"] += 1
    print(f"    ! {req_id} failed after {MAX_RETRIES} tries "
          f"({last}); recorded as no-flags")
    from shared.evaluator import CRITERIA_ORDER
    return {"violated": {cid: False for cid in CRITERIA_ORDER}, "evaluations": []}


def _run_once(reqs, context, sme_rules, provider) -> Dict[str, Dict]:
    """One full pass over all requirements -> {req_id: analyze result}.

    `reqs` is a list of (req_id, text) drawn from the Project Specification."""
    results: Dict[str, Dict] = {}
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futs = {pool.submit(_analyze_one, rid, txt, context, sme_rules, provider): rid
                for rid, txt in reqs}
        for fut in as_completed(futs):
            rid = futs[fut]
            results[rid] = fut.result()
    return results


def _matrix_sheet(ws, ds, results) -> None:
    ws.append(["Rule ID", "Rule Description"] + ds.req_ids)
    for cid in CRITERIA:
        row = [cid, CRITERIA_NAMES[cid]]
        for rid in ds.req_ids:
            violated = results.get(rid, {}).get("violated", {}).get(cid, False)
            row.append("x" if violated else None)
        ws.append(row)


def _gt_workbook(ds) -> None:
    """Write the ground-truth matrix once per dataset, same shape as predictions."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GT"
    ws.append(["Rule ID", "Rule Description"] + ds.req_ids)
    for cid in CRITERIA:
        row = [cid, CRITERIA_NAMES[cid]]
        for rid in ds.req_ids:
            row.append("x" if ds.gt[cid][rid] else None)
        ws.append(row)
    wb.save(RUNS_DIR / f"{ds.name}_gt.xlsx")


def run(dataset: str, config: str, n_runs: int, provider: str = "anthropic") -> Path:
    ds = datasets.load_dataset(dataset)          # GT + req_id ordering (scoring)
    spec_in = spec.load_spec(dataset)            # INPUT: context + requirement text
    context = spec_in.context
    sme_rules = inject.load_sme_rules(config)

    # Analyze the spec's canonical requirement text, in GT column order.
    reqs = [(rid, spec_in.requirements.get(rid, "")) for rid in ds.req_ids]
    missing = [rid for rid, txt in reqs if not txt]
    if missing:
        raise ValueError(f"{dataset}: spec missing text for {missing}")

    OUTPUT_DIR.mkdir(exist_ok=True)
    _gt_workbook(ds)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    raw_all: Dict[str, Dict] = {}

    for run_idx in range(1, n_runs + 1):
        t0 = time.time()
        results = _run_once(reqs, context, sme_rules, provider)
        ws = wb.create_sheet(f"Execution_{run_idx:02d}")
        _matrix_sheet(ws, ds, results)
        raw_all[f"Execution_{run_idx:02d}"] = {
            rid: results[rid]["evaluations"] for rid in results}
        n_flags = sum(sum(r["violated"].values()) for r in results.values())
        print(f"  {dataset}/{config} run {run_idx:02d}/{n_runs}: "
              f"{n_flags} flags, {time.time()-t0:.0f}s")

    pred_path = RUNS_DIR / f"{dataset}_{config}_matrix.xlsx"
    wb.save(pred_path)
    (RUNS_DIR / f"{dataset}_{config}_raw.json").write_text(
        json.dumps(raw_all, indent=2, ensure_ascii=False), encoding="utf-8")
    return pred_path


def main() -> None:
    ap = argparse.ArgumentParser(description="Run prediction matrices")
    ap.add_argument("--runs", type=int, default=10)
    ap.add_argument("--datasets", nargs="+", default=["PM", "Bookshelf"])
    ap.add_argument("--configs", nargs="+", default=["baseline", "M1", "M2"])
    ap.add_argument("--provider", default="anthropic")
    args = ap.parse_args()

    print(f"Running {args.runs} runs x {args.datasets} x {args.configs} "
          f"on {args.provider}")
    for dataset in args.datasets:
        for config in args.configs:
            path = run(dataset, config, args.runs, args.provider)
            print(f"  -> {path.name}")
    print(f"\nDone. Matrices + GT + raw evaluations in {RUNS_DIR}")
    if _DEGRADED["n"]:
        print(f"WARNING: {_DEGRADED['n']} cell(s) recorded as no-flags after "
              f"exhausting retries.")


if __name__ == "__main__":
    main()
