"""KPI engine — score prediction matrices against ground truth.

Matches the confusion-metric definitions used in the AI4RE evaluation notebook /
colab_R1_R14_evaluation script:

    accuracy    = (TP+TN)/(TP+TN+FP+FN)
    precision   = TP/(TP+FP)
    recall      = TP/(TP+FN)      (sensitivity)
    specificity = TN/(TN+FP)
    F1          = 2*P*R/(P+R)
    FPR         = FP/(FP+TN)
    FNR         = FN/(FN+TP)

For each (dataset, config) it computes, per execution:
  * per-rule confusion + metrics (across requirements)
  * a POOLED (micro-averaged) metric over the whole matrix — the headline number

then aggregates across the N executions (mean +/- std), writes a KPI workbook,
and per-rule boxplot PNGs.
"""

from __future__ import annotations

import math
from pathlib import Path
from typing import Dict, List, Optional

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd

from shared.paths import TESTBED_RUNS_DIR as RUNS_DIR, output_dir

# Matrices, raw dumps and ground-truth workbooks are the RUN STORE: each one
# cost 100 whole-set LLM calls, so they are versioned rather than rebuilt.
# Plots and summary workbooks are regenerable and go to outputs/.
RUNS_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR = output_dir("n8n_testbed")
PLOT_DIR = OUTPUT_DIR / "plots"

PLOT_METRICS = ["accuracy", "precision", "recall", "f1", "fpr", "fnr"]


def _safe_div(n: float, d: float) -> Optional[float]:
    return n / d if d else None


def _metrics(tp: int, tn: int, fp: int, fn: int) -> Dict[str, Optional[float]]:
    # Detection metrics (precision/recall/F1) are only meaningful when the unit
    # actually has ground-truth positives to detect. A criterion with zero true
    # violations (e.g. PM's Necessary/Appropriate) has TP+FN=0: recall is 0/0 and
    # precision collapses to 0 the instant the model flags it, which is a
    # degenerate score, not a model quality. Report those as N/A (None) so they
    # drop out of precision/recall/F1 plots. FPR/specificity/accuracy remain
    # defined (they measure the false-alarm side, which IS meaningful here).
    has_positives = (tp + fn) > 0
    if has_positives:
        prec = _safe_div(tp, tp + fp)
        rec = _safe_div(tp, tp + fn)
        f1 = _safe_div(2 * prec * rec, prec + rec) if (prec and rec) else (
            0.0 if (prec == 0 or rec == 0) else None)
    else:
        prec = rec = f1 = None
    return {
        "tp": tp, "tn": tn, "fp": fp, "fn": fn,
        "accuracy": _safe_div(tp + tn, tp + tn + fp + fn),
        "precision": prec,
        "recall": rec,
        "specificity": _safe_div(tn, tn + fp),
        "f1": f1,
        "fpr": _safe_div(fp, fp + tn),
        "fnr": _safe_div(fn, fn + tp),
    }


def _read_matrix(path: Path, sheet: str) -> Dict[str, Dict[str, bool]]:
    """{rule_id: {req_id: flagged?}} from a matrix sheet."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    req_cols = {i: str(h) for i, h in enumerate(header) if i >= 2 and h}
    out: Dict[str, Dict[str, bool]] = {}
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        rid = str(row[0]).strip()
        out[rid] = {req_cols[i]: (str(row[i]).strip().lower() == "x")
                    for i in req_cols}
    return out


def _gt_map(dataset: str) -> Dict[str, Dict[str, bool]]:
    """Ground truth for scoring, honouring AI4RE_GT_SOURCE.

    A source-tagged workbook (PM_gt_survey.xlsx) wins when present, so the two
    ground truths can coexist and a run can be scored against either without
    overwriting the other. Falls back to the untagged file.
    """
    from shared.datasets import gt_source
    tagged = RUNS_DIR / f"{dataset}_gt_{gt_source()}.xlsx"
    path = tagged if tagged.exists() else RUNS_DIR / f"{dataset}_gt.xlsx"
    return _read_matrix(path, "GT")


def score(dataset: str, config: str, criteria=None) -> pd.DataFrame:
    """Long per-(execution, rule) metrics table for one (dataset, config).

    `criteria` restricts BOTH the per-rule rows and the POOLED confusion to a
    subset. Needed whenever two arms do not cover the same criteria: a criterion
    absent from one side's matrix scores as "never flagged" rather than "not
    evaluated", so its ground-truth positives silently become false negatives
    and depress that arm's recall.
    """
    gt = _gt_map(dataset)
    if criteria is not None:
        keep = set(criteria)
        missing = keep - set(gt)
        if missing:
            raise ValueError(f"no ground-truth row for {sorted(missing)}")
        gt = {k: v for k, v in gt.items() if k in keep}
    pred_path = RUNS_DIR / f"{dataset}_{config}_matrix.xlsx"
    wb = openpyxl.load_workbook(pred_path, data_only=True)

    records: List[Dict] = []
    for sheet in wb.sheetnames:
        pred = _read_matrix(pred_path, sheet)
        pooled = {"tp": 0, "tn": 0, "fp": 0, "fn": 0}
        for rid, gcol in gt.items():
            pcol = pred.get(rid, {})
            tp = tn = fp = fn = 0
            for req, g in gcol.items():
                p = pcol.get(req, False)
                tp += p and g
                tn += (not p) and (not g)
                fp += p and (not g)
                fn += (not p) and g
            for k, v in zip(("tp", "tn", "fp", "fn"), (tp, tn, fp, fn)):
                pooled[k] += v
            rec = {"execution": sheet, "rule": rid, **_metrics(tp, tn, fp, fn)}
            records.append(rec)
        records.append({"execution": sheet, "rule": "POOLED",
                        **_metrics(pooled["tp"], pooled["tn"],
                                   pooled["fp"], pooled["fn"])})
    df = pd.DataFrame(records)
    df.insert(0, "config", config)
    df.insert(0, "dataset", dataset)
    return df


def summarize(df: pd.DataFrame) -> pd.DataFrame:
    """Mean +/- std of each metric per rule, across executions."""
    metric_cols = ["accuracy", "precision", "recall", "specificity", "f1", "fpr", "fnr"]
    g = df.groupby(["dataset", "config", "rule"])[metric_cols]
    summary = g.agg(["mean", "std"])
    summary.columns = [f"{m}_{s}" for m, s in summary.columns]
    return summary.reset_index()


def _boxplot_by_rule(df: pd.DataFrame, metric: str, dataset: str, config: str,
                     path: Path) -> None:
    rules = [r for r in df["rule"].unique() if r != "POOLED"] + ["POOLED"]
    data = [df[df["rule"] == r][metric].dropna().values for r in rules]
    if not any(len(d) for d in data):
        return
    fig, ax = plt.subplots(figsize=(9, 4.5))
    ax.boxplot(data, tick_labels=rules, patch_artist=True,
               boxprops=dict(facecolor="#8a1538", alpha=0.6),
               medianprops=dict(color="black"))
    ax.set_title(f"{dataset} / {config} — {metric} by rule (n={df['execution'].nunique()} runs)")
    ax.set_ylabel(metric)
    ax.set_ylim(-0.05, 1.05)
    ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    fig.savefig(path, dpi=120)
    plt.close(fig)


def run(datasets: List[str], configs: List[str]) -> pd.DataFrame:
    PLOT_DIR.mkdir(parents=True, exist_ok=True)
    all_long = []
    for dataset in datasets:
        for config in configs:
            pred_path = RUNS_DIR / f"{dataset}_{config}_matrix.xlsx"
            if not pred_path.exists():
                print(f"  skip {dataset}/{config} (no matrix)")
                continue
            df = score(dataset, config)
            all_long.append(df)
            for metric in PLOT_METRICS:
                _boxplot_by_rule(df, metric, dataset, config,
                                 PLOT_DIR / f"{dataset}_{config}_{metric}.png")
            pooled = df[df["rule"] == "POOLED"]
            print(f"  {dataset}/{config}: pooled "
                  f"acc={pooled['accuracy'].mean():.3f} "
                  f"prec={pooled['precision'].mean():.3f} "
                  f"rec={pooled['recall'].mean():.3f} "
                  f"f1={pooled['f1'].mean():.3f}")
    long_df = pd.concat(all_long, ignore_index=True)
    summary = summarize(long_df)

    with pd.ExcelWriter(OUTPUT_DIR / "kpi_metrics.xlsx") as xl:
        long_df.to_excel(xl, sheet_name="per_execution", index=False)
        summary.to_excel(xl, sheet_name="summary", index=False)
    return summary


def main() -> None:
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--datasets", nargs="+", default=["PM", "Bookshelf"])
    ap.add_argument("--configs", nargs="+", default=["baseline", "M1", "M2"])
    args = ap.parse_args()
    run(args.datasets, args.configs)
    print("\nWrote kpi_metrics.xlsx and per-rule plots in outputs/plots/")


if __name__ == "__main__":
    main()
