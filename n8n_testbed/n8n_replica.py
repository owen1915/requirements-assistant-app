"""Replicate the n8n 'AI4RE GR_subrules' assessment PROCESS in the testbed.

Fidelity items closed (see the --rubric / --model / --provider flags):

  * WHOLE SET in one prompt   — all requirements evaluated in a single call,
    exactly like the n8n agent (not one call per requirement).
  * n8n RUBRIC WORDING        — --rubric n8n loads incose_rules_n8n.json, whose
    A4.3 / A9.2 / A9.3 carry allowances the prototype rubric omits. This is the
    single biggest driver of the A4 and A9 numbers.
  * EXPLANATIONS REQUESTED    — the model returns criterion + name + a 1-2
    sentence explanation per violation, exactly as n8n asks. The prose is then
    discarded at matrix-write time (the n8n Transform node does the same) but
    kept in the raw dump. Asking a model to justify a flag changes what it is
    willing to flag, so this is part of the process, not an extra.
  * FR -> PR -> ER -> RR ORDER — ports n8n's sortRequirementIds. Position
    matters in a whole-set call.
  * A1 NORMALIZATION          — if A1 ("no issues") is returned for a
    requirement, its other A-flags are cleared (ports the Transform code node).
  * A1 / A11 RECORDED         — both stay out of the scored matrix (the GT has
    no rows for them) but are persisted to the raw dump and reported as an
    A11-diversion count and an A1-suppression count. Both are pure recall loss
    that was previously invisible.
  * NO context-leniency clause — the testbed's "if context supplies the need,
    don't flag" sentence is intentionally omitted (n8n has none).

Retries are deliberately NOT removed: n8n retries nothing and turns a parse
failure into an all-blank matrix that scores as "found no issues", which is a
defect rather than a behaviour worth reproducing.

Output matrices are written in the same Execution_NN shape the KPI engine reads,
so metrics.py / notebook_eval.py / boxplots.py score it with no changes.

    python -m n8n_testbed.n8n_replica --runs 10 --datasets PM --rubric n8n
"""

from __future__ import annotations

import argparse
import json
import os
import re
import time
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl

from shared.llm import ensure_env
ensure_env()

from shared import datasets, spec
from . import n8n_prompt_verbatim as VERBATIM
from shared.datasets import CRITERIA, CRITERIA_NAMES
from .inject import (EVAL_MODEL, TEMPERATURE, _CALL_TIMEOUT, _strip_to_json,
                     load_sme_rules)

from shared.paths import TESTBED_RUNS_DIR as RUNS_DIR, output_dir

# Matrices, raw dumps and ground-truth workbooks are the RUN STORE: each one
# cost 100 whole-set LLM calls, so they are versioned rather than rebuilt.
# Plots and summary workbooks are regenerable and go to outputs/.
RUNS_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR = output_dir("n8n_testbed")
MAX_RETRIES = 4


# Config -> output tag. "baseline" injects no rules; the others inject the ICAI
# rule groups into the same whole-set prompt so the process is held constant.
def _config_tag(config: str) -> str:
    return "n8nreplica" if config == "baseline" else f"n8n_{config}"


# Rubric id -> file. "incose" is the prototype's wording; "n8n" is the
# workflow's own, which is more permissive on A4.3 / A9.2 / A9.3.
from shared.paths import RUBRIC_DIR
RUBRICS = {
    "incose": RUBRIC_DIR / "incose_rules.json",
    "n8n": RUBRIC_DIR / "incose_rules_n8n.json",
}


def _all_a_criteria(rubric: str) -> List[Dict]:
    return json.loads(RUBRICS[rubric].read_text(encoding="utf-8"))["individual_criteria"]


def _criteria_block(sme_rules: Dict[str, List[str]], rubric: str) -> str:
    lines = []
    for c in _all_a_criteria(rubric):
        cid = c["criterion_id"]
        lines.append(f"**{cid} - {c['name']}**: {c['description']}")
        for sr in c.get("sub_rules", []):
            lines.append(f"  - {sr}")
        for learned in sme_rules.get(cid, []):
            lines.append(f"  - [Expert-reviewed guidance from prior reviews] {learned}")
    return "\n".join(lines)


def _build_system(sme_rules: Dict[str, List[str]], rubric: str) -> str:
    """The n8n PART-1 (individual A-criteria) framing, minus the B block."""
    criteria = _criteria_block(sme_rules, rubric)
    example = (
        '{"individualEvaluations": {'
        '"FR.1": [{"criterion": "A3", "name": "Appropriate", "explanation": "..."}, '
        '{"criterion": "A5", "name": "Complete", "explanation": "..."}], '
        '"FR.2": [{"criterion": "A1", "name": "No Issues", '
        '"explanation": "No issues found with this requirement."}]}}'
    )
    return f"""You are a requirements quality evaluator. You will receive a list of requirements with their IDs, along with Project Context and Concept of Operations (ConOps) information.

When completing the following task, carefully consider the provided Project Context and Concept of Operations to determine if requirements align with the project goals, needs, and operational constraints.

Evaluate EACH requirement individually against ALL of these quality criteria (A1-A11). Be CRITICAL and THOROUGH in identifying issues.

IMPORTANT: Criterion A1 means NO violation of any of the criteria for a requirement. Criteria A2-A11 mean a violation of that criterion EXISTS. For each requirement, include ALL the criteria that are violated.

Be thorough and critical during the evaluation. If violations exist, return one object per criterion, including the criterion ID, the criterion name, and a short explanation of why the violation exists (1-2 sentences).

For each criterion, use both the criterion DESCRIPTION and the SUB-RULES as a checklist to guide your judgment. If the description or any sub-rule is violated, the criterion is violated. Evaluate each criterion independently for each requirement. Where expert-reviewed guidance is given under a criterion, weight it - it reflects how expert reviewers judged similar cases.

{criteria}

OUTPUT FORMAT - return ONLY valid JSON, no markdown, no preamble. Start with {{ and end with }}. Map every requirement ID to the list of its violations. Use a single A1 object if the requirement has no issues:
{example}"""


def _build_user(context: str, reqs: List) -> str:
    body = "\n".join(f"{rid}: {txt}" for rid, txt in reqs)
    ctx = context.strip() if context else "No additional context provided."
    return (f"Project Context and Concept of Operations:\n{ctx}\n\n"
            f"Requirements to evaluate:\n{body}")


# --- n8n sortRequirementIds: FR -> PR -> ER -> RR, then numeric ---------------
_PREFIX_ORDER = {"FR": 0, "PR": 1, "ER": 2, "RR": 3}


def _sort_req_ids(ids: List[str]) -> List[str]:
    def key(rid: str):
        prefix = (rid.split(".")[0] or "").upper()
        try:
            num = int(rid.split(".")[1])
        except (IndexError, ValueError):
            num = 0
        return (_PREFIX_ORDER.get(prefix, 999), num)
    return sorted(ids, key=key)


def _call_set(system: str, user: str, model: str, provider: str) -> str:
    if provider == "anthropic":
        import anthropic
        key = os.getenv("ANTHROPIC_API_KEY", "").strip()
        client = anthropic.Anthropic(api_key=key, timeout=_CALL_TIMEOUT, max_retries=0)
        resp = client.messages.create(
            model=model, max_tokens=8000, temperature=TEMPERATURE,
            system=[{"type": "text", "text": system,
                     "cache_control": {"type": "ephemeral"}}],
            messages=[{"role": "user", "content": user}])
        return resp.content[0].text.strip()

    if provider == "openrouter":
        # The workflow file itself runs openai/gpt-5.5 through OpenRouter. Kept
        # behind a flag so the model is a deliberate choice, not an accident.
        import openai as openai_lib
        client = openai_lib.OpenAI(
            api_key=os.getenv("OPENROUTER_API_KEY", "").strip(),
            base_url="https://openrouter.ai/api/v1", timeout=_CALL_TIMEOUT)
        resp = client.chat.completions.create(
            model=model, max_tokens=8000, temperature=TEMPERATURE,
            messages=[{"role": "system", "content": system},
                      {"role": "user", "content": user}])
        return resp.choices[0].message.content.strip()

    raise ValueError(f"unknown provider {provider!r}")


def _criterion_ids(entries) -> List[str]:
    """Violation objects (or bare ids) -> list of criterion ids."""
    out = []
    for e in entries if isinstance(entries, list) else []:
        cid = e.get("criterion") or e.get("criterion_id") or "" if isinstance(e, dict) else e
        cid = str(cid).strip().upper()
        if re.fullmatch(r"A\d{1,2}", cid):
            out.append(cid)
    return out


def _normalize_a1(ids: List[str]) -> List[str]:
    """Port of the n8n normalizeAIssues: if A1 is present, keep only A1."""
    return ["A1"] if "A1" in ids else ids


def _evaluate_once(system: str, user: str, model: str,
                   provider: str) -> Tuple[Dict, Dict, Dict]:
    """One whole-set call -> (raw ids, A1-normalized ids, full objects)."""
    last = None
    for attempt in range(MAX_RETRIES):
        try:
            raw_text = _call_set(system, user, model, provider)
            data = _strip_to_json(raw_text)
            indiv = data.get("individualEvaluations", {})
            objs = {rid: (v if isinstance(v, list) else [])
                    for rid, v in indiv.items()}
            raw = {rid: _criterion_ids(v) for rid, v in objs.items()}
            norm = {rid: _normalize_a1(ids) for rid, ids in raw.items()}
            return raw, norm, objs
        except Exception as exc:  # noqa: BLE001
            last = exc
            time.sleep(2 * (attempt + 1))
    raise RuntimeError(f"whole-set call failed after {MAX_RETRIES} tries: {last}")


def _matrix_sheet(ws, ds, per_req: Dict[str, List[str]]) -> None:
    ws.append(["Rule ID", "Rule Description"] + ds.req_ids)
    for cid in CRITERIA:                       # scored A2-A10 rows
        row = [cid, CRITERIA_NAMES[cid]]
        for rid in ds.req_ids:
            flagged = cid in per_req.get(rid, [])
            row.append("x" if flagged else None)
        ws.append(row)


def _gt_workbook(ds) -> None:
    wb = openpyxl.Workbook(); w = wb.active; w.title = "GT"
    w.append(["Rule ID", "Rule Description"] + ds.req_ids)
    for cid in CRITERIA:
        w.append([cid, CRITERIA_NAMES[cid]] +
                 ["x" if ds.gt[cid][rid] else None for rid in ds.req_ids])
    wb.save(RUNS_DIR / f"{ds.name}_gt.xlsx")


def _diagnostics(ds, raw: Dict, norm: Dict) -> Dict[str, int]:
    """A1/A11 recall loss that never reaches the scored matrix."""
    suppressed = 0     # true flags erased by the A1 normalization
    diverted = 0       # requirements routed to A11 that GT says have a real flag
    for rid in ds.req_ids:
        raw_ids, norm_ids = raw.get(rid, []), norm.get(rid, [])
        if "A1" in raw_ids:
            suppressed += sum(1 for c in raw_ids
                              if c in CRITERIA and ds.gt[c][rid])
        if "A11" in norm_ids and any(ds.gt[c][rid] for c in CRITERIA):
            diverted += 1
    return {"a1_suppressed_true_flags": suppressed, "a11_diverted_reqs": diverted}


def run(dataset: str, n_runs: int, config: str = "baseline",
        rubric: str = "n8n", model: str = EVAL_MODEL,
        provider: str = "anthropic", prompt: str = "verbatim") -> Path:
    ds = datasets.load_dataset(dataset)
    sp = spec.load_spec(dataset)

    # Prompt order is SHEET ROW ORDER (what Process Sheet Data emits), not the
    # sorted order — sortRequirementIds only orders the matrix columns.
    scored = set(ds.req_ids)
    reqs = [(rid, txt) for rid, txt in sp.requirements.items() if rid in scored]
    missing = [rid for rid in ds.req_ids if not sp.requirements.get(rid)]
    if missing:
        raise ValueError(f"{dataset}: spec missing text for {missing}")

    sme_rules = load_sme_rules(config)          # {} for baseline
    if prompt == "verbatim":
        if sme_rules:
            raise ValueError("--prompt verbatim cannot carry injected SME rules; "
                             "use --prompt generated for non-baseline configs")
        system = VERBATIM.SYSTEM_PART1
        user = VERBATIM.build_user(sp.project_context, sp.conops, reqs,
                                   sp.image_url)
    else:
        system = _build_system(sme_rules, rubric)
        user = _build_user(sp.context, reqs)
    tag = _config_tag(config)
    n_rules = sum(len(v) for v in sme_rules.values())

    OUTPUT_DIR.mkdir(exist_ok=True)
    _gt_workbook(ds)

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    raw_dump = []
    totals = {"a1_suppressed_true_flags": 0, "a11_diverted_reqs": 0}
    for run_idx in range(1, n_runs + 1):
        t0 = time.time()
        raw, norm, objs = _evaluate_once(system, user, model, provider)
        ws = wb.create_sheet(f"Execution_{run_idx:02d}")
        _matrix_sheet(ws, ds, norm)

        diag = _diagnostics(ds, raw, norm)
        for k, v in diag.items():
            totals[k] += v
        raw_dump.append({"run": run_idx, "raw": raw, "normalized": norm,
                         "evaluations": objs, "diagnostics": diag})

        n_flags = sum(len([c for c in v if c not in ("A1", "A11")])
                      for v in norm.values())
        print(f"  {dataset}/{tag} ({n_rules} rules) run {run_idx:02d}/{n_runs}: "
              f"{n_flags} A-flags, A1-suppressed {diag['a1_suppressed_true_flags']}, "
              f"A11-diverted {diag['a11_diverted_reqs']}, {time.time() - t0:.0f}s")

    path = RUNS_DIR / f"{dataset}_{tag}_matrix.xlsx"
    wb.save(path)
    (RUNS_DIR / f"{dataset}_{tag}_raw.json").write_text(
        json.dumps({"dataset": dataset, "config": config, "rubric": rubric,
                    "model": model, "provider": provider,
                    "prompt": prompt, "temperature": TEMPERATURE,
                    "runs": n_runs, "req_order": [r for r, _ in reqs],
                    "totals": totals,
                    "executions": raw_dump}, indent=2),
        encoding="utf-8")
    print(f"  totals over {n_runs} runs: {totals}")
    return path


def main() -> None:
    ap = argparse.ArgumentParser(description="n8n process replica (A-criteria)")
    ap.add_argument("--runs", type=int, default=15)
    ap.add_argument("--datasets", nargs="+", default=["PM"])
    ap.add_argument("--configs", nargs="+", default=["baseline"],
                    help="baseline | lenient | default | moderate | strict")
    ap.add_argument("--rubric", choices=sorted(RUBRICS), default="n8n",
                    help="n8n = the workflow's own sub-rule wording (default)")
    ap.add_argument("--model", default=EVAL_MODEL)
    ap.add_argument("--provider", choices=["anthropic", "openrouter"],
                    default="anthropic")
    ap.add_argument("--prompt", choices=["verbatim", "generated"],
                    default="verbatim",
                    help="verbatim = the workflow's PART 1 text, transcribed")
    args = ap.parse_args()
    print(f"n8n replica: {args.runs} runs x {args.datasets} x {args.configs} "
          f"on {args.model} ({args.provider}), prompt={args.prompt}, "
          f"rubric={args.rubric if args.prompt == 'generated' else 'n/a (verbatim)'}")
    tags = []
    for dataset in args.datasets:
        for config in args.configs:
            path = run(dataset, args.runs, config, args.rubric, args.model,
                       args.provider, args.prompt)
            tags.append(_config_tag(config))
            print(f"  -> {path.name}")
    print("\nDone. Score with:  python -m n8n_testbed.metrics "
          f"--datasets {' '.join(args.datasets)} --configs {' '.join(sorted(set(tags)))}")


if __name__ == "__main__":
    main()
