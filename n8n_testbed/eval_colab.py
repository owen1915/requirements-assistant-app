"""Score our prediction matrices with the ACTUAL AI4RE notebook logic.

Instead of reimplementing the metrics, this converts our A-series prediction
matrices to the R-series the notebook uses and drives colab_R1_R14_evaluation_v2
directly (run_metrics -> run_visualization -> run_per_rule_analysis), scoring
against the notebook's own GT file (R1_R14_GroundTruth.xlsx).

    python -m n8n_testbed.eval_colab --configs M1 M2
"""

from __future__ import annotations

import argparse
import importlib.util
from pathlib import Path

import matplotlib
matplotlib.use("Agg")   # set before the colab module imports pyplot
import openpyxl

from shared.paths import TESTBED_RUNS_DIR as RUNS_DIR, output_dir

# Matrices, raw dumps and ground-truth workbooks are the RUN STORE: each one
# cost 100 whole-set LLM calls, so they are versioned rather than rebuilt.
# Plots and summary workbooks are regenerable and go to outputs/.
RUNS_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR = output_dir("n8n_testbed")
from shared.paths import GT_DIR, REPO_ROOT

# The AI4RE evaluation script is driven as-is rather than reimplemented, so it
# is vendored alongside the other reference material instead of being read out
# of whichever folder it happened to be downloaded to.
COLAB_PATH = (REPO_ROOT / "reference" / "ai4re_notebook"
              / "colab_R1_R14_evaluation_v2.py")
GT_PATH = GT_DIR / "R1_R14_GroundTruth.xlsx"
GT_SHEET = "Sheet1"

# prototype criterion (A-series) -> notebook rule id (R-series)
A_TO_R = {"A2": "R1", "A3": "R12", "A4": "R3", "A5": "R4",
          "A6": "R2", "A9": "R5", "A10": "R11"}


def _to_rseries(src: Path, dst: Path) -> None:
    """Copy a prediction workbook, relabelling the Rule ID column A->R."""
    wb = openpyxl.load_workbook(src)
    for ws in wb.worksheets:
        for r in range(2, ws.max_row + 1):          # row 1 = header
            v = ws.cell(r, 1).value
            key = str(v).strip() if v is not None else ""
            if key in A_TO_R:
                ws.cell(r, 1).value = A_TO_R[key]
    wb.save(dst)


def _load_colab():
    if not COLAB_PATH.exists():
        raise FileNotFoundError(
            f"the AI4RE evaluation script is missing at {COLAB_PATH}; it is "
            "vendored in reference/ai4re_notebook/")
    spec = importlib.util.spec_from_file_location("colab_ev", str(COLAB_PATH))
    mod = importlib.util.module_from_spec(spec)
    mod.__name__ = "colab_ev"
    spec.loader.exec_module(mod)
    return mod


def run(configs) -> None:
    colab = _load_colab()
    for config in configs:
        pred_a = RUNS_DIR / f"PM_{config}_matrix.xlsx"
        if not pred_a.exists():
            print(f"skip {config}: no {pred_a.name}")
            continue
        pred_r = RUNS_DIR / f"PM_{config}_matrix_rseries.xlsx"
        _to_rseries(pred_a, pred_r)

        confusion = RUNS_DIR / f"PM_{config}_confusion.xlsx"
        summary = RUNS_DIR / f"PM_{config}_summary.xlsx"
        boxplot = OUTPUT_DIR / f"PM_{config}_notebook_boxplot.png"

        print(f"\n{'='*66}\nPM / {config}  (notebook logic)\n{'='*66}")
        colab.run_metrics(str(pred_r), str(GT_PATH), GT_SHEET, str(confusion))
        colab.run_visualization(str(confusion), str(summary), str(boxplot))
        colab.run_per_rule_analysis(str(confusion))


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--configs", nargs="+", default=["M1", "M2"])
    args = ap.parse_args()
    run(args.configs)
    print(f"\nOutputs in {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
